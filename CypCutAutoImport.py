import time, os.path
import win32process, win32api
import openpyxl
from openpyxl import   Workbook, load_workbook
from pywinauto import Desktop, Application, mouse, keyboard, findwindows
import tkinter as tk
from tkinter import filedialog as fd
from tkinter import Tk, ttk, Button, Label, messagebox


GUI = tk.Tk()
GUI.geometry('400x170')
GUI.title('CypCut AutoImport')
GUI.resizable(False, False)


def shielding(String):

    return '\"' + String + '\"' + ' '


def Initial_File_Button_click(event):

    filetypes = (
                 ('xlsx', '*.xlsx'),
                 ('Все типы', '*.*')
                 )
    global Initial_filename
    Initial_filename = fd.askopenfilename(
                title='Открыть файл Excel',
                initialdir='/',
                filetypes=filetypes)

    Initial_Label = ttk.Label(GUI, text=Initial_filename)
    Initial_Label.grid(column=1, row=0, sticky=tk.SW, padx=0, pady=0)


def Start_Button_click(event):


    try:
        CypCutHandle = findwindows.find_window(best_match='CypCut')
        app = Application(backend="uia").connect(handle=CypCutHandle)
        appWnd = app.top_window()
        CypCut = Desktop()[appWnd.texts()[0]]
        ThreatID,ProcessID = win32process.GetWindowThreadProcessId(CypCutHandle)
        layoutTID = win32api.GetKeyboardLayout(ThreatID)
    except:
        Message1 = tk.messagebox.showwarning('CypCut не запущен!', 'Запустите CypCut')
        return

    if layoutTID !=67699721:
         Message3 = tk.messagebox.showwarning('Язык ввода: Русский', 'Выберете английскую раскладку для CypCut')
         return

    try:
        Wb = load_workbook(Initial_filename)
        Ws = Wb.worksheets[0]
    except :
        Message2 = tk.messagebox.showwarning('Ошибка ввода', 'Выберете Excel файл')
        return

    if CypCut.is_minimized():
        CypCut.maximize()

    CypCut.set_focus()
    
    # Окно Import Parts

    Tapp = Desktop()['Import Parts']
    Edit = Tapp.child_window(class_name="ComboBoxEx32").child_window(class_name="Edit")
    OpenButton = Tapp.child_window(title="&Открыть", class_name="Button")
    PartsList = CypCut.child_window(class_name="TFrmNest").child_window(class_name="TFsCheckListBox")

    for i in range(0, Ws.max_row):

        # Заход в меню File->Nest->ImportParts
        # Для хоткеев надо поставить vk_packet=False

        keyboard.send_keys('{VK_MENU}'+'f', pause=0.1, vk_packet=False)
        keyboard.send_keys('p', pause=0.1, vk_packet=False)
        keyboard.send_keys('a', pause=0.1, vk_packet=False)
        time.sleep(0.3)
        Edit.set_text(Ws.cell(i + 1, 1).value)
        OpenButton.click()
        time.sleep(0.5)

        # Выбор последнего добавленного элемента
        # Нулевой индекс принадлежит пустому элементу, поэтому -1

        PartsList.select(PartsList.item_count() - 1)
        PartItem = PartsList.child_window(class_name='TRzCalcEdit')
        PartItem.set_text(Ws.cell(i + 1, 2).value)
        time.sleep(0.5)


Initial_File_Button = ttk.Button(GUI, text='Excel файл')
Initial_File_Button.grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)

Start_Button = ttk.Button(GUI, text='Импортировать')
Start_Button.grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)

Initial_File_Button.bind("<Button-1>", Initial_File_Button_click)
Start_Button.bind("<Button-1>", Start_Button_click)

GUI.mainloop()
